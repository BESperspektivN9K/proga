import openpyxl
from docx import Document
import re
import math
from openpyxl.styles import PatternFill

# Открываем .docx
doc = Document(r"C:\Users\User\Documents\диплом\Возможно поможет в написании диплома(1)\парам.docx")
KTGS=1.0
# Создаем новую Excel-книгу и два листа
wb = openpyxl.Workbook()
ws_t0 = wb.active
ws_t0.title = "T0"
ws_20t0 = wb.create_sheet(title="20T0")
ws_errors =wb.create_sheet(title="Праметры с ошибкой") 

# Заголовки
headers1 = ["Наименование", "Идентификатор", "Номер 16р слова", "Начальный участок выдачи", "Конечный участок выдачи", "Тип", "Задача/Блок", "Количество слов"]
headers2 = ["Наименование", "Идентификатор", "Номер 16р слова","Номер счетчика", "Начальный участок выдачи", "Конечный участок выдачи", "Тип", "Задача/Блок", "Количество слов"]
ws_t0.append(headers1)
ws_20t0.append(headers2)
ws_errors.append(headers1)

# Массивы
words = []
used_words = []
rows_T0 = []
rows_20T0 = []
#rows_with_r = []
rows_errors = []
available_time = [100,101,105,110,120,125,130,135,140,145,150,155,160,165,170,175,180,185,190,195,200,205,210,215,220,225,230,240,255]
available_tip = [1,2,3,4,5,6,7]
number_word = 51

# --- Первый проход: читаем строки и определяем, куда записывать ---
for i, row in enumerate(doc.tables[0].rows):
    if i == 0:
        continue

    cells = [cell.text.strip() for cell in row.cells]
    if len(cells) < 5:
        continue

    name = cells[0]
    id = cells[1]
    group = cells[2]  # Новый столбец: T0 или 20T0

    numbers = re.findall(r"\d+", cells[3])
    if len(numbers) == 3:
        start_num = numbers[1]
        end_num = numbers[2]
    else:
        start_num = numbers[0]
        end_num = numbers[1]

    words_to_add = 1.0
    """""
    has_r = 'р' in cells[-2]

    if has_r:
        match = re.search(r'(\d+)\s*р', cells[-2])
        words_to_add = float(match.group(1)) / 16.0 if match else 1.0
    else:
        numbers = re.findall(r'(\d+\.?\d*)', cells[-2].split('/')[0])  # числа до "/"
        for num in numbers:
            words_to_add *= float(num)#нормальная проверка 
    # Перемножаем все числа до "/"
    """
    if 'р' in cells[-2]:
        match = re.search(r'(\d+)\s*р', cells[-2])
        words_to_add = float(match.group(1)) / 16.0
    else:
        numbers = re.findall(r'(\d+\.?\d*)', cells[-2].split('/')[0])  # числа до "/"
        for num in numbers:
            words_to_add *= float(num)
    # Перемножаем все числа до "/"
    # Проверка на наличие "Ктгс" и умножение на KTGS
    if 'Ктгс' in cells[-2]:
        words_to_add *= KTGS
    
    tip_match = re.findall(r"/(\d+)", cells[-2])
    tip = tip_match[0] if tip_match else ""
    tip = tip[0] if tip else ""

    block = cells[-1]

    # Выбираем нужный лист
    target_ws = ws_t0 if group == "T0" else ws_20t0
    excel_row = target_ws.max_row + 1

    row_data = {
        "name": name,
        "id": id,
        "group": group,
        "start_num": start_num,
        "end_num": end_num,
        "tip": tip,
        "block": block,
        "words_to_add": words_to_add,
        "excel_row": excel_row
    }

    #if has_r and group == "T0":
    #    rows_with_r.append(row_data)
    #else:
    #    rows_T0.append(row_data)

    if int(start_num) not in available_time or int(end_num) not in available_time or int(tip) not in available_tip or group not in ['T0', '20T0'] or (int(start_num) >= int(end_num)):
        rows_errors.append(row_data)
    else:
        if group == "20T0":
            rows_20T0.append(row_data)
            target_ws.append([name, id, "","", start_num, end_num, tip, block, words_to_add])
        else:
            rows_T0.append(row_data)
            target_ws.append([name, id, "", start_num, end_num, tip, block, words_to_add])
    

# --- Сортировка и выдача слов ---
rows_20T0.sort(key=lambda row: row["words_to_add"], reverse=True)

total_words_20t0 = sum(int(row["words_to_add"]) for row in rows_20T0)  # Сколько слов нужно всего
words_per_counter = 20  # Максимальное количество счётчиков, которое можно использовать

# Вычисляем количество слов, которые можем выдать для каждого счётчика
count_words = 2 if total_words_20t0 < words_per_counter else math.ceil(total_words_20t0 / words_per_counter)

# --- Переменные для учёта слов ---
available_words = [str(i + 1) for i in range(count_words)]  # Список доступных слов
word_index = 0
counter = 1  # Начальный счётчик

final_rows_20T0 = []  # Новый список для хранения строк с учётом счётчиков

i = 0
while i < len(rows_20T0):
    row = rows_20T0[i]
    remaining = int(row["words_to_add"])  # Сколько слов осталось для текущей строки
    name, id, start_num, end_num, tip, block = row["name"], row["id"], row["start_num"], row["end_num"], row["tip"], row["block"]
    
    # --- Выдаём слова с учётом счётчиков ---
    while remaining > 0:
        # Сколько слов можем взять в этом счётчике
        take = min(remaining, count_words)  # В зависимости от оставшихся слов и максимума на счётчик
        used = [available_words[j % count_words] for j in range(take)]  # Выбираем слова по счётчику

        # Формируем диапазон слов
        word_range = f"{int(used[0])+number_word -1}" if len(used) == 1 else f"{int(used[0])+number_word -1}-{int(used[-1])+number_word -1}"

        # Добавляем строку с текущим счётчиком
        final_rows_20T0.append({
            "name": name,
            "id": id,
            "word_range": word_range,
            "counter": counter,
            "start_num": start_num,
            "end_num": end_num,
            "tip": tip,
            "block": block,
            "count": take
        })

        remaining -= take  # Уменьшаем оставшееся количество слов
        counter += 1  # Увеличиваем счётчик

    i += 1

# --- Очистка листа и перезапись заголовков ---
ws_20t0.delete_rows(2, ws_20t0.max_row)  # Очищаем старые строки, оставляем только заголовки

# --- Записываем строки на лист 20T0 ---
for row in final_rows_20T0:
    ws_20t0.append([ 
        row["name"],
        row["id"],
        row["word_range"],
        row["counter"],
        row["start_num"],
        row["end_num"],
        row["tip"],
        row["block"],
        row["count"]
    ])
number_word += count_words
# --- Обработка строк для T0 ---

for row in rows_T0:
    start_num = row["start_num"]
    end_num = row["end_num"]
    words_to_add = row["words_to_add"]
    excel_row = row["excel_row"]  # Получаем номер строки для листа T0
    first_word = number_word

    for _ in range(int(words_to_add)):
        words.append({
            "номер": str(number_word),
            "начальный": start_num,
            "конечный": end_num
        })
        used_words.append({
            "номер": str(number_word),
            "начальный": start_num,
            "конечный": end_num
        })
        number_word += 1

    last_word = number_word - 1
    word_range = f"{first_word}" if first_word == last_word else f"{first_word}-{last_word}"

    ws_t0.cell(row=excel_row, column=3).value = word_range

for row in rows_T0:#переделано 
    start_num = row["start_num"]
    end_num = row["end_num"]
    excel_row = row["excel_row"]
    group = row["group"]
    words_to_add = int(row["words_to_add"])
    target_ws = ws_t0 if group == "T0" else ws_20t0
    assigned_words = []

    # Пробуем найти подряд идущие свободные слова
    for i in range(len(used_words) - words_to_add + 1):
        candidate_group = used_words[i:i + words_to_add]
        if all(int(word["конечный"]) < int(start_num) for word in candidate_group):
            # Назначаем слова
            for word in candidate_group:
                word["конечный"] = end_num
                for w in words:
                    if w["номер"] == word["номер"]:
                        w["конечный"] = end_num
                        break
                assigned_words.append(word["номер"])
            break  # Нашли и назначили — выходим из цикла

    # Записываем в Excel диапазон, если слова были найдены
    if assigned_words:
        word_range = assigned_words[0] if len(assigned_words) == 1 else f"{assigned_words[0]}-{assigned_words[-1]}"
        target_ws.cell(row=excel_row, column=3).value = word_range
        
red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid") 
for row in rows_errors:
    ws_errors.append([
    row["name"],
    row["id"],
    "",
    row["start_num"],
    row["end_num"],
    row["tip"],
    row["block"],
    row["words_to_add"]
])
    last_row = ws_errors.max_row
    for col in range(1, 9):  # Подкрашиваем ячейки A-H
        ws_errors.cell(row=last_row, column=col).fill = red_fill
# --- Сохраняем ---
wb.save(r"C:\Users\User\Documents\диплом\результат.xlsx")
print("Файл сохранен!")
# --- Вывод слов ---
print("\nМассив слов:")
for word in sorted(words, key=lambda w: int(w["номер"])):
    print(f"Слово №{word['номер']}: {word['начальный']} → {word['конечный']}")

