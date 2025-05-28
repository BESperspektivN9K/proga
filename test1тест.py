import openpyxl
from docx import Document
import re
import math
from openpyxl.styles import PatternFill

def first_word_num(word_range):
    if '-' in word_range:
        return int(word_range.split('-')[0])
    else:
        return int(word_range)
    
# Открываем .docx
doc = Document(r"C:\Users\User\Documents\диплом\парам.docx")
KTGS=1.0
# Создаем новую Excel-книгу и два листа
wb = openpyxl.Workbook()
ws_t0 = wb.active
ws_t0.title = "T0"
ws_20t0 = wb.create_sheet(title="20T0")
ws_errors =wb.create_sheet(title="Праметры с ошибкой") 

# Заголовки
headers1 = ["Наименование", "Идентификатор", "Номер 16р слова", "Начальный участок выдачи", "Конечный участок выдачи", "Тип", "Номер разряда", "Задача/Блок", "Количество слов"]
headers2 = ["Наименование", "Идентификатор", "Номер 16р слова","Номер счетчика", "Начальный участок выдачи", "Конечный участок выдачи", "Тип", "Задача/Блок", "Количество слов"]
ws_t0.append(headers1)
ws_20t0.append(headers2)
ws_errors.append(headers1)

# Массивы
words = []
used_words = []
rows_T0 = []
rows_20T0 = []
rows_T0_raz = []
excel_rowT0 = 1
excel_row_raz = 1
rows_errors = []
available_time = [100,101,105,110,120,125,130,135,140,145,150,155,160,165,170,175,180,185,190,195,200,205,210,215,220,225,230,240,255]
available_tip = [1,2,3,4,5,6,7,8]
number_word = 51


# --- Первый проход: читаем строки и определяем, куда записывать ---
for i, row in enumerate(doc.tables[0].rows):
    if i == 0:
        continue

    cells = [cell.text.strip() for cell in row.cells]
    if len(cells) < 5:
        continue

    name = cells[0]
    id = cells[1]
    group = cells[2]  # Новый столбец: T0 или 20T0

    numbers = re.findall(r"\d+", cells[3])
    if len(numbers) == 3:
        start_num = numbers[1]
        end_num = numbers[2]
    else:
        start_num = numbers[0]
        end_num = numbers[1]

    
    words_to_add = 1.0
    if 'р' in cells[-2]:  
        match = re.findall(r'(\d+)р', cells[-2])
        kolvo = float(match[-1])
        words_to_add = float(kolvo) / 16.0
    elif ',' in cells[-2]:
        normal_value = cells[-2].replace(',','.')
        kolvo = normal_value.split('/')[0] 
        numbers = re.findall(r'\d+\.?\d*',kolvo)
        for num in numbers:
            words_to_add *= float(num)
    else:
        kolvo = cells[-2].split('/')[0]
        numbers = re.findall(r'(\d+)', kolvo)
    # Перемножаем все числа до "/"
        for num in numbers:
            words_to_add *= float(num)
    
    # Проверка на наличие "Ктгс" и умножение на KTGS
    if 'Ктгс' in cells[-2]:
        words_to_add *= KTGS
    
    tip_match = re.findall(r"/(\d+)", cells[-2])
    tip = tip_match[0] if tip_match else "8"
    tip = tip[0] if tip else "8"

    block = cells[-1]

    # Выбираем нужный лист
    if group == "T0":
        if (isinstance(words_to_add, float) and words_to_add != int(words_to_add)):
            excel_row_raz += 1
            excel_row = excel_row_raz
        else:
            target_ws = ws_t0
            excel_rowT0 += 1
            excel_row = excel_rowT0
    else:
        target_ws = ws_20t0
        excel_row = target_ws.max_row + 1


    row_data = {
        "name": name,
        "id": id,
        "group": group,
        "start_num": start_num,
        "end_num": end_num,
        "tip": tip,
        "block": block,
        "words_to_add": words_to_add,
        "excel_row": excel_row
    }

    if int(start_num) not in available_time or int(end_num) not in available_time or int(tip) not in available_tip or group not in ['T0', '20T0'] or (int(start_num) >= int(end_num)):
        rows_errors.append(row_data)
    else:
        if group == "20T0":
            rows_20T0.append(row_data)
            target_ws.append([name, id, "","", start_num, end_num, tip, block, words_to_add])
        else:
            if (isinstance(words_to_add, float) and words_to_add != int(words_to_add)):
                rows_T0_raz.append(row_data)
            else:
                rows_T0.append(row_data)
    
"""""
    if (id == previous_id) and (end_num == previous_start or start_num == previous_end):
        if end_num == previous_start:
            # previous_start = start_num
            rows_T0[-2][3] = start_num
            # ws_t0.cell(row=excel_row - 1, column=4).value = previous_start
        else:
            # previous_end = end_num
            rows_T0[-2][4] = end_num
            # ws_t0.cell(row=excel_row-1, column=5).value = previous_end
        excel_rowT0 -= 1
    else:
    """""

for row in rows_T0:
    ws_t0.append([row["name"],
        row["id"],
        "",
        row["start_num"],
        row["end_num"],
        row["tip"],
        "",
        row["block"],
        row["words_to_add"]
        ])

# --- Сортировка и выдача слов ---
#rows_20T0.sort(key=lambda row: row["words_to_add"], reverse=True)
rows_T0.sort(key=lambda row: row["start_num"], reverse=False)


words_per_counter = 20
total_words_20t0 = sum(int(row["words_to_add"]) for row in rows_20T0)
count_words = 2 if total_words_20t0 < words_per_counter else math.ceil(total_words_20t0 / words_per_counter)

# Инициализация
used_words_20t0 = []
counter = 1
final_rows_20T0 = []

max_counters = 20
count_words_per_counter = count_words
available_words = [str(number_word + i) for i in range(count_words_per_counter)]

used_words_per_counter = {i: [] for i in range(1, max_counters + 1)}

final_rows_20T0 = []

for row in rows_20T0:
    name = row["name"]
    id = row["id"]
    start_num = int(row["start_num"])
    end_num = int(row["end_num"])
    tip = row["tip"]
    block = row["block"]
    words_needed = int(row["words_to_add"])

    counter = 1

    while words_needed > 0 and counter <= max_counters:
        used_in_counter = used_words_per_counter[counter]

        free_words = []
        for word in available_words:
            conflict = False
            for used in used_in_counter:
                if word == used["номер"]:
                    if not (end_num <= used["start"] or start_num >= used["end"]):
                        conflict = True
                        break
            if not conflict:
                free_words.append(word)

        if not free_words:
            counter += 1
            continue

        take = min(words_needed, len(free_words))
        taken_words = free_words[:take]

        # Помечаем слова занятыми для текущего счетчика
        for word in taken_words:
            used_words_per_counter[counter].append({
                "номер": word,
                "start": start_num,
                "end": end_num
            })

        # Формируем диапазон слов для этой части
        word_range = taken_words[0] if len(taken_words) == 1 else f"{taken_words[0]}-{taken_words[-1]}"

        final_rows_20T0.append({
            "name": name,
            "id": id,
            "word_range": word_range,
            "counter": counter,
            "start_num": start_num,
            "end_num": end_num,
            "tip": tip,
            "block": block,
            "count": take
        })

        words_needed -= take
        counter += 1  # Переходим к следующему счетчику для остатка слов

    if words_needed > 0:
        raise RuntimeError(f"Не удалось назначить все слова для параметра: {name}")
# --- Очистка листа и перезапись заголовков ---
ws_20t0.delete_rows(2, ws_20t0.max_row)  # Очищаем старые строки, оставляем только заголовки

final_rows_20T0.sort(key=lambda x: (x["counter"], first_word_num(x["word_range"])))
# --- Записываем строки на лист 20T0 ---
for row in final_rows_20T0:
    ws_20t0.append([ 
        row["name"],
        row["id"],
        row["word_range"],
        row["counter"],
        row["start_num"],
        row["end_num"],
        row["tip"],
        row["block"],
        row["count"]
    ])
number_word += count_words

max_counter = max(row["counter"] for row in final_rows_20T0)
ws_t0.merge_cells('J3:Q3')
ws_t0.merge_cells('J4:Q4')
ws_t0['J3'] = f"{'Количество слов, выделенное под параметры 20Т0:'} {count_words} {'('} " \
              f"{number_word-count_words} {'-'} {number_word-1} {')'}"
ws_t0['J4'] = f"Количество счетчиков, использованных для распределения параметров 20Т0:  {max_counter}"

# --- Обработка строк для T0 ---


for row in rows_T0:
    start_num = row["start_num"]
    end_num = row["end_num"]
    words_to_add = row["words_to_add"]
    excel_row = row["excel_row"]  # Получаем номер строки для листа T0
    first_word = number_word

    for _ in range(int(words_to_add)):
        words.append({
            "номер": str(number_word),
            "начальный": start_num,
            "конечный": end_num
        })
        used_words.append({
            "номер": str(number_word),
            "начальный": start_num,
            "конечный": end_num
        })
        number_word += 1

    last_word = number_word - 1
    word_range = f"{first_word}" if first_word == last_word else f"{first_word}-{last_word}"

    ws_t0.cell(row=excel_row, column=3).value = word_range

for row in rows_T0:
    start_num = row["start_num"]
    end_num = row["end_num"]
    excel_row = row["excel_row"]
    group = row["group"]
    words_to_add = row["words_to_add"]                                       #row["words_to_add"] * row["count"]
    target_ws = ws_t0 if group == "T0" else ws_20t0
    assigned_words = []

    for _ in range(int(words_to_add)):
        for word in used_words:
            if int(word["конечный"]) < int(start_num):
                word["конечный"] = end_num
                for w in words:
                    if w["номер"] == word["номер"]:
                        w["конечный"] = end_num
                        break
                assigned_words.append(word["номер"])
                break

    if assigned_words:
        word_range = assigned_words[0] if len(assigned_words) == 1 else f"{assigned_words[0]}-{assigned_words[-1]}"
        target_ws.cell(row=excel_row, column=3).value = word_range

red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid") 
for row in rows_errors:
    ws_errors.append([
    row["name"],
    row["id"],
    "",
    row["start_num"],
    row["end_num"],
    row["tip"],
    row["block"],
    row["words_to_add"]
])
    last_row = ws_errors.max_row
    for col in range(1, 9):  # Подкрашиваем ячейки A-H
        ws_errors.cell(row=last_row, column=col).fill = red_fill

sum_raz = 0
counter_raz_word = 1
counter_raz = 0
rows_T0_raz.sort(key=lambda row: row["start_num"], reverse=False)
for row in rows_T0_raz:#разрядные отдельно добавляю
    ws_t0.append([row["name"],
        row["id"],
        "",
        row["start_num"],
        row["end_num"],
        row["tip"] if row["tip"] == '7' else "",
        "",
        row["block"],
        row["words_to_add"]
        ])
    row["excel_row"] = ws_t0.max_row 
    sum_raz += row["words_to_add"]

b = 0
k = 1
raz_words = math.ceil(sum_raz)
while b < len(rows_T0_raz):# чет здесь не работает
    row = rows_T0_raz[b]
    razrad = row["words_to_add"] * 16
    if k + razrad > 17:
        number_word += 1
        k = 1
    else:
        ws_t0.cell(row=row["excel_row"], column=3).value = number_word
        if razrad > 1:
            ws_t0.cell(row=row["excel_row"], column=7).value = f"{k}-{k+razrad-1}р"
        else:
            ws_t0.cell(row=row["excel_row"], column=7).value = f"{k}р"
        k += razrad
        b += 1


# --- Сохраняем ---
wb.save(r"C:\Users\User\Documents\диплом\Параметры.xlsx")
print("Файл сохранен!")
# --- Вывод слов ---
print("\nМассив слов:")
for word in sorted(words, key=lambda w: int(w["номер"])):
    print(f"Слово №{word['номер']}: {word['начальный']} → {word['конечный']}")

